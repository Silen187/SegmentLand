import numpy as np
import pwlf
import matplotlib
matplotlib.use('Agg') 
import matplotlib.pyplot as plt
import math
import pandas as pd
import xarray as xr
from openpyxl import load_workbook
from dask import delayed
from dask import compute
from dask.distributed import Client




np.random.seed(42)

class ChangeDetection:

    def __init__(self, params, data, min_year):
        """
        Khởi tạo lớp ChangeDetection với đầu vào từ người dùng.
        """
        # Nhập các tham số từ người dùng, với giá trị mặc định
        self.n_segments = params["n_segments"]
        self.spike_threshold = params["spike_threshold"]
        self.vertex_count_overshoot = params["vertex_count_overshoot"]
        self.recovery_threshold = params["recovery_threshold"]
        self.preventOneYearRecovery = params["preventOneYearRecovery"]
        self.minObservation = params["minObservation"]

        self.data = data
        self.min_year = min_year
        self.smoothed_data = None
        self.smoothed_data_normalized = None
        self.breakpoints = None
        self.slopes = None
        self.pwlf_model = None

    @staticmethod
    def calculate_angle_between_slopes(m1, m2):
        """
        Tính góc giữa hai đoạn thẳng dựa trên độ dốc.
        """
        cos_theta = abs((1 + m1 * m2) / (math.sqrt(1 + m1**2) * math.sqrt(1 + m2**2)))
        return math.degrees(math.acos(cos_theta))


    def calculate_rmse(self):
        """
        Tính RMSE (Root Mean Square Error) của mô hình LandTrendr.
        """
        y_true = self.smoothed_data
        y_pred = self.pwlf_model.predict(np.arange(len(y_true)))
        rmse = np.sqrt(np.mean((y_true - y_pred) ** 2))
        return rmse
    
    @staticmethod
    def calculate_variance(y_true, y_pred):
        """
        Tính phương sai giữa giá trị thực và giá trị dự đoán.
        """
        return np.mean((y_true - y_pred) ** 2)

    def adjust_duplicates(self, lst):
        lst = [round(num) for num in lst]
        max_value = lst[-1]  # Giá trị lớn nhất ban đầu trong danh sách
        i = 0
        while i < len(lst) - 1:
            # Nếu phần tử hiện tại trùng với phần tử tiếp theo
            if lst[i] == lst[i + 1]:
                lst[i] += 1  # Cộng thêm 1 cho phần tử hiện tại
                
                # Kiểm tra nếu giá trị mới lớn hơn max_value, loại bỏ phần tử
                if lst[i] > max_value:
                    lst.pop(i)
                    i -= 1  # Điều chỉnh chỉ số để không bỏ qua phần tử nào
                else:
                    lst.sort()
                    i -= 1
            i += 1
        return lst


    def smooth_spikes(self):
        """
        Làm mịn dữ liệu để loại bỏ nhiễu bằng cách lấy trung bình của các điểm liền kề khi phát hiện nhiễu.
        """
        nbr_values = self.data
        smoothed_values = nbr_values.copy()
        for i in range(1, len(nbr_values) - 1):
            if nbr_values[i] != 0:
                spike_ratio_before = abs(nbr_values[i] - nbr_values[i-1]) / abs(nbr_values[i-1])
                spike_ratio_after = abs(nbr_values[i+1] - nbr_values[i]) / abs(nbr_values[i+1])
                if spike_ratio_before >= self.spike_threshold and spike_ratio_after >= self.spike_threshold:
                    smoothed_values[i] = (nbr_values[i-1] + nbr_values[i+1]) / 2
        self.smoothed_data = smoothed_values

    def segment_with_overshoot(self):
        """
        Thực hiện hồi quy phân đoạn với số lượng đoạn tối đa được phép và thêm overshoot.
        """
        # Chuẩn hóa smoothed_data
        self.smoothed_data_normalized = self.smoothed_data  / (self.smoothed_data.max() - self.smoothed_data.min()) * (len(self.smoothed_data) - 1)
        
        # Thực hiện hồi quy phân đoạn với dữ liệu đã chuẩn hóa
        time_numeric = np.arange(len(self.smoothed_data_normalized))
        my_pwlf = pwlf.PiecewiseLinFit(time_numeric, self.smoothed_data_normalized)
        breakpoints = my_pwlf.fit(self.n_segments + self.vertex_count_overshoot)
        self.breakpoints = self.adjust_duplicates(breakpoints)
        my_pwlf.fit_with_breaks(self.breakpoints)
        self.slopes = my_pwlf.slopes
        self.pwlf_model = my_pwlf

    def dynamic_thresholds(self, slopes, nbr_values, nbr_percentile):
        """
        Tính toán các ngưỡng động dựa trên độ dốc và sự thay đổi NBR.
        """
        result = {}
        for i in range(len(slopes) - 1):
            theta = self.calculate_angle_between_slopes(slopes[i], slopes[i+1])
            result[i] = {'slope': slopes[i], 'slope_next': slopes[i+1], 'theta': theta}
        
        nbr_diff = np.abs(np.diff(nbr_values))
        nbr_threshold = np.percentile(nbr_diff, nbr_percentile)
        
        return result, nbr_threshold
    
    def filter_breakpoints_dynamic(self, breakpoints, slopes, nbr_values):
        """
        Lọc các điểm phân đoạn dựa trên ngưỡng động.
        """
        nbr_percentile = max(20, 70 - 10 * self.n_segments * (np.std(nbr_values) / (0.5*(len(nbr_values)-1))))
        
        nbr_percentile = min(nbr_percentile, 60)

        slope_threshold_result, nbr_threshold = self.dynamic_thresholds(slopes, nbr_values, nbr_percentile)
        
        filtered_breakpoints = [breakpoints[0]]  # Giữ lại điểm đầu tiên
        
        for i in range(1, len(breakpoints) - 1):
            delta_nbr_after = abs(nbr_values[breakpoints[i]+1] - self.pwlf_model.predict([breakpoints[i]])[0])
            delta_nbr_before = abs(nbr_values[breakpoints[i]-1] - self.pwlf_model.predict([breakpoints[i]])[0])
            
            # Kiểm tra các điều kiện dựa trên ngưỡng động
            if (delta_nbr_after > nbr_threshold or delta_nbr_before > nbr_threshold) and \
               (((slope_threshold_result[i-1]['slope'] * slope_threshold_result[i-1]['slope_next']) < 0) or \
                ((slope_threshold_result[i-1]['slope'] * slope_threshold_result[i-1]['slope_next']) >= 0 and \
                 slope_threshold_result[i-1]['theta'] >= 30)):
                filtered_breakpoints.append(breakpoints[i])
        
        # Giữ lại điểm cuối cùng
        filtered_breakpoints.append(breakpoints[-1])
        
        self.breakpoints = filtered_breakpoints

    def remove_breakpoint_with_max_variance(self, time_numeric, nbr_values, max_segments):
        """
        Loại bỏ đỉnh gây ra sự gia tăng phương sai lớn nhất bằng cách sử dụng pwlf.
        """
        iteration = 1  # Biến để đếm số lần lặp
        while len(self.breakpoints) > max_segments + 1:  # Cộng 1 vì giữ điểm đầu và cuối
            variances = []
            temp_pwlf = []
            # Duyệt qua tất cả các đỉnh, bỏ từng đỉnh một và tính phương sai
            for i in range(1, len(self.breakpoints) - 1):  # Bỏ qua đỉnh đầu và cuối
                # Tạo danh sách breakpoints sau khi loại bỏ đỉnh thứ i
                temp_breakpoints = self.breakpoints[:i] + self.breakpoints[i+1:]
                
                # Sử dụng pwlf để thực hiện hồi quy phân đoạn với breakpoints tạm thời
                my_pwlf = pwlf.PiecewiseLinFit(time_numeric, nbr_values)
                my_pwlf.fit_with_breaks(temp_breakpoints)
                
                temp_pwlf.append(my_pwlf)
                # Dự đoán giá trị NBR sau khi hồi quy
                y_pred = my_pwlf.predict(time_numeric)

                # Tính phương sai giữa giá trị thực và giá trị dự đoán
                total_variance = self.calculate_variance(nbr_values, y_pred)
                
                # Lưu lại phương sai khi loại đỉnh thứ i
                variances.append(total_variance)
            
            # Tìm đỉnh có phương sai lớn nhất và loại bỏ
            max_variance_idx = np.argmin(variances) + 1  # Cộng 1 vì bỏ qua đỉnh đầu
            self.pwlf_model = temp_pwlf[np.argmin(variances)]
            # Loại bỏ đỉnh này khỏi danh sách breakpoints
            del self.breakpoints[max_variance_idx]
            # Tăng biến đếm lần lặp
            iteration += 1
        self.slopes = self.pwlf_model.slopes

    def detect_and_remove_short_term_recovery(self):
        """
        Loại bỏ các phục hồi ngắn hạn (1 năm) trong danh sách breakpoints.
        """
        filtered_breakpoints = [self.breakpoints[0]]  # Giữ lại điểm đầu tiên
        
        for i in range(1, len(self.breakpoints) - 1):
            # Tính thời gian giữa các điểm breakpoints
            duration = self.breakpoints[i + 1] - self.breakpoints[i]
            recovery_speed = (self.pwlf_model.predict(self.breakpoints[i + 1]) - self.pwlf_model.predict(self.breakpoints[i])) / \
                             (self.pwlf_model.predict(self.breakpoints[i - 1]) - self.pwlf_model.predict(self.breakpoints[i]))
            
            # Phát hiện phục hồi ngắn hạn
            if (self.slopes[i - 1] < 0 and self.slopes[i] > 0) and (0.9 <= recovery_speed <= 1.2) and duration <= 1:
                continue  # Bỏ qua điểm phục hồi ngắn hạn này
            
            filtered_breakpoints.append(self.breakpoints[i])
        
        # Giữ lại điểm cuối cùng
        filtered_breakpoints.append(self.breakpoints[-1])
        
        self.breakpoints = filtered_breakpoints  # Cập nhật breakpoints đã loại bỏ phục hồi ngắn hạn

    def detect_and_remove_small_recoveries(self):
        """
        Loại bỏ các phục hồi không đáng kể trong danh sách breakpoints.
        
        Parameters:
        - recovery_threshold: Ngưỡng phục hồi tối thiểu để được coi là đáng kể.
        """
        recovery_threshold = self.recovery_threshold

        filtered_breakpoints = [self.breakpoints[0]]  # Giữ lại điểm đầu tiên
        
        for i in range(1, len(self.breakpoints) - 1):
            if self.slopes[i - 1] < 0 and self.slopes[i] > 0:  # Nếu đây là phục hồi
                duration = self.breakpoints[i + 1] - self.breakpoints[i]
                recovery_magnitude = ((self.pwlf_model.predict(self.breakpoints[i + 1])[0] - self.pwlf_model.predict(self.breakpoints[i])[0]) / duration) / \
                                     (self.pwlf_model.predict(self.breakpoints[i - 1])[0] - self.pwlf_model.predict(self.breakpoints[i])[0])
                
                if not (duration >= 2 and recovery_magnitude > recovery_threshold):
                    filtered_breakpoints.append(self.breakpoints[i])
            else:
                filtered_breakpoints.append(self.breakpoints[i])
        
        # Giữ lại điểm cuối cùng
        filtered_breakpoints.append(self.breakpoints[-1])
        
        self.breakpoints = filtered_breakpoints
    
    def final_model(self):
        time_numeric = np.arange(len(self.smoothed_data))
        my_pwlf = pwlf.PiecewiseLinFit(time_numeric, self.smoothed_data)
        my_pwlf.fit_with_breaks(self.breakpoints)
        self.pwlf_model = my_pwlf

    def plot_result(self, pixel_id):
        """
        Vẽ đồ thị kết quả cuối cùng với các breakpoints đã lọc.
        """
        time_numeric = np.arange(len(self.smoothed_data))
        plt.figure(figsize=(12, 6))
        plt.plot(time_numeric, self.smoothed_data, 'o', label='Smoothed Data')
        plt.plot(time_numeric, self.data, 'v', label='Origin Data')
                
        for i in range(1, len(self.breakpoints)):
            x_segment = np.linspace(self.breakpoints[i-1], self.breakpoints[i], num=100)
            y_segment = self.pwlf_model.predict(x_segment)
            plt.plot(x_segment, y_segment, '-', color='red')
        
        for bp in self.breakpoints:
            plt.axvline(x=bp, linestyle='--', color='green', label=f'Breakpoint at {bp:.2f}')
        
        plt.title(f"Final Piecewise Linear Regression with Filtered Breakpoints For Pixel {pixel_id}")
        plt.xlabel('Time')
        plt.ylabel('NBR')
        plt.legend()
        plt.grid(True)
        plt.savefig(f"pixel_{pixel_id}.png")
        plt.close()  # Đóng biểu đồ để giải phóng bộ nhớ


    def get_segment_data(self, pixel_id, delta='all', filename="all_pixels_segments.xlsx"):
        """
        Tạo mảng thông tin về các phân đoạn dựa trên các điểm breakpoint, giữ nguyên hướng delta.
        
        Parameters:
        - delta (String): 'all', 'loss', hoặc 'gain' để chọn loại phân đoạn.
        
        Returns:
        - Một mảng 2D với 11 hàng, mỗi cột đại diện cho một phân đoạn.
        """
        segments_info = []

        for i in range(len(self.breakpoints) - 1):
            start = self.breakpoints[i]
            end = self.breakpoints[i + 1]
            start_value_predict = self.pwlf_model.predict([start])[0]
            end_value_predict = self.pwlf_model.predict([end])[0]
            delta_value_predict = end_value_predict - start_value_predict
            start_value = self.smoothed_data[self.breakpoints[i]]
            end_value = self.smoothed_data[self.breakpoints[i+1]]
            duration = end - start
            rate_of_change_predict = delta_value_predict / duration 
            
            # Lọc các phân đoạn dựa trên loại delta ('all', 'loss', 'gain')
            if (delta == 'all') or (delta == 'loss' and delta_value_predict < 0) or (delta == 'gain' and delta_value_predict > 0):
                segments_info.append([
                    start + self.min_year,               # Hàng 1: Năm bắt đầu
                    end + self.min_year,                 # Hàng 2: Năm kết thúc
                    start_value_predict,         # Hàng 3: Giá trị bắt đầu
                    end_value_predict,           # Hàng 4: Giá trị kết thúc
                    delta_value_predict,        #Giá trị thay đổi quang phổ dự đoán
                    start_value,            #Giá trị làm trơn thực bắt đầu
                    end_value,              #Giá trị làm trơn kết thúc
                    duration,            # Hàng 6: Thời lượng thay đổi
                    rate_of_change_predict,     # Hàng 7: Tốc độ thay đổi quang phổ dự đoán
                    delta_value_predict / self.calculate_rmse()  # Hàng 8: DSNR (chuẩn hóa theo RMSE)
                ])

        columns = [
            'Năm Bắt Đầu', 'Năm Kết Thúc', 'Giá Trị Dự Đoán Bắt Đầu', 'Giá Trị Dự Đoán Kết Thúc', 'Sự Thay Đổi Quang Phổ Dự Đoán',
            'Giá Trị Thực Bắt Đầu', 'Giá Trị Thực Kết Thúc',
            'Thời Lượng', 'Tốc Độ Thay Đổi Dự Đoán', 'DSNR'
        ]

        df = pd.DataFrame(segments_info, columns=columns).T

        start_row = pixel_id * 11  # Mỗi pixel chiếm 11 hàng và cách nhau 1 hàng
        
        try:
            with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                # Ghi nhãn pixel_id tại hàng bắt đầu
                sheet = writer.book['Sheet1']
                sheet.cell(row=start_row + 1, column=1, value=f"Pixel ID: {pixel_id}")
                
                # Ghi dữ liệu của DataFrame từ hàng tiếp theo
                df.to_excel(writer, sheet_name="Sheet1", startrow=start_row, startcol=1, index=True, header=False)
        
        except FileNotFoundError:
            # Nếu file chưa tồn tại, tạo file mới
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Ghi nhãn pixel_id và dữ liệu vào file Excel
                sheet = writer.sheets.get("Sheet1", writer.book.create_sheet("Sheet1"))
                sheet.cell(row=start_row + 1, column=1, value=f"Pixel ID: {pixel_id}")
                df.to_excel(writer, sheet_name="Sheet1", startrow=start_row, startcol=1, index=True, header=False)


@delayed
def process_pixel_dask(pixel_id, params, data, min_year):
    # Khởi tạo một đối tượng ChangeDetection mới cho mỗi pixel
    detector = ChangeDetection(params, data, min_year)
    # Kiểm tra nếu dữ liệu không đủ số lượng quan sát
    if detector.data is not None and len(detector.data) < detector.minObservation:
        raise ValueError(f"Dữ liệu không đủ số lượng quan sát tối thiểu: yêu cầu {detector.minObservation}, nhưng chỉ có {len(detector.data)}.")
    # Tiến hành các bước phân tích
    detector.smooth_spikes()
    detector.segment_with_overshoot()
    detector.filter_breakpoints_dynamic(detector.breakpoints, detector.slopes, detector.smoothed_data_normalized)
    detector.remove_breakpoint_with_max_variance(np.arange(len(detector.smoothed_data_normalized)), detector.smoothed_data_normalized, detector.n_segments)
    
    if detector.preventOneYearRecovery:
        detector.detect_and_remove_short_term_recovery()
    
    detector.detect_and_remove_small_recoveries()
    detector.final_model()
    detector.get_segment_data(pixel_id)
    detector.plot_result(pixel_id)


def run_parallel_with_client(params, all_data, min_year):
    # Khởi tạo Dask Client
    client = Client(n_workers=8, threads_per_worker=2, memory_limit='2GB')

    # Tạo danh sách các tác vụ delayed cho tất cả các pixel
    tasks = [process_pixel_dask(pixel_id.item(), params, all_data.NBR.sel(pixel_id=pixel_id).values, min_year) for pixel_id in all_data.pixel_id]
    print(len(tasks))
    
    # Chạy song song và theo dõi tiến trình
    compute(*tasks)  # Dask sẽ tự động phân phối công việc cho các lõi CPU

    client.close()  # Đóng client khi xong
    print("Thành công!")

def import_data(file_path):

    # Đọc tệp CSV vào DataFrame
    data = pd.read_csv(file_path)  
    years = [int(col.split('_')[1]) for col in data.columns if col.startswith("Year_")]
    pixel_ids = data['pixel_id']

    ds = xr.Dataset(
        {
            "NBR": (["pixel_id", "year"], data.drop(columns="pixel_id").values)
        },
        coords={
            "pixel_id": pixel_ids,
            "time": years
        }
    )

    return ds, min(years)

params = {
    "n_segments": 6,
    "spike_threshold": 0.7,
    "vertex_count_overshoot": 4,
    "recovery_threshold": 0.25,
    "preventOneYearRecovery": True,
    "minObservation": 6
}

if __name__ == '__main__':
    all_data, min_year = import_data('pivot_data.csv')
    run_parallel_with_client(params, all_data, min_year)